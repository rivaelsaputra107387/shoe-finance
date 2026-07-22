"""
Script to extract COA and Journal data from the Excel file.
Run this script to generate PHP seeder-ready data.

Usage: python extract_excel_data.py
"""
import openpyxl
import json

wb = openpyxl.load_workbook('LAPORAN KEUANGAN SHOE WORKSHOP (1).xlsx', data_only=True)
print("=== Sheet Names ===")
print(wb.sheetnames)

# Extract COA data
print("\n=== COA Sheet ===")
if 'COA' in wb.sheetnames:
    ws = wb['COA']
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), values_only=False):
        values = [cell.value for cell in row]
        if any(v is not None for v in values):
            print(values)

# Extract CONTROL data
print("\n=== CONTROL Sheet ===")
if 'CONTROL' in wb.sheetnames:
    ws = wb['CONTROL']
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100), values_only=False):
        values = [cell.value for cell in row]
        if any(v is not None for v in values):
            print(values)

# Extract JURNAL data (first 30 rows for sample)
print("\n=== JURNAL Sheet ===")
if 'JURNAL' in wb.sheetnames:
    ws = wb['JURNAL']
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=False):
        values = [cell.value for cell in row]
        if any(v is not None for v in values):
            print(values)

# Extract BUKU BESAR sample
print("\n=== BUKU BESAR Sheet ===")
if 'BUKU BESAR' in wb.sheetnames:
    ws = wb['BUKU BESAR']
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=False):
        values = [cell.value for cell in row]
        if any(v is not None for v in values):
            print(values)

print("\n=== Done ===")
